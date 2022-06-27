from bs4 import BeautifulSoup
import requests
import openpyxl
from datetime import date,datetime

url = requests.get("https://pt.surebet.com/calculator/show/aAxmn2gWc94")

conteudo = url.content
soup =BeautifulSoup(conteudo, 'html.parser')
soup.encode= "utf-8"


# Coleta de dados

dataAtual = date.today()
dataPlanilha = dataAtual.strftime('%d/%m/%y')
horaAtual = datetime.now()
horaPlanilha =horaAtual.strftime('%H:%M')

jogo = soup.find("span", attrs={"class": "calc_event"}).text

infoC1 = soup.find("tr", attrs={"data-number": "0"})
nomeC1 = infoC1.find("td", attrs={"class": "booker_c"}).text
oddC1 = infoC1.find("input", attrs={"class": "koefficient form-control inline-input w-number"})

infoC2 = soup.find("tr", attrs={"data-number": "1"})
nomeC2 = infoC2.find("td", attrs={"class": "booker_c"}).text
oddC2 = infoC2.find("input", attrs={"class": "koefficient form-control inline-input w-number"})
nomeC2.encode("utf8")
nomeC1.encode("utf8")


caminho = "D:\Repositorio_de_Estudo\Pessoal\SureBet.xlsx"
planilha = openpyxl.load_workbook(caminho)
pagina = planilha['Planilha de Apostas']

#acha a proxima linha vazia
vazia = 17
for row in pagina.iter_rows(min_row=17):
    if(row[4].value == None):
        break
    vazia += 1

#Joga dados na planilha
pagina.cell(row=vazia, column=4).value = dataPlanilha
pagina.cell(row=vazia, column=5).value = nomeC1
pagina.cell(row=vazia+1, column=5).value = nomeC2
pagina.cell(row=vazia, column=6).value = horaPlanilha
pagina.cell(row=vazia, column=7).value = jogo
pagina.cell(row=vazia, column=9).value = float(oddC1['value'])
pagina.cell(row=vazia+1, column=9).value = float(oddC2['value'])
pagina.cell(row=vazia, column=10).value = 2.000
pagina.cell(row=vazia+1, column=10).value = 2.000

planilha.save(caminho)